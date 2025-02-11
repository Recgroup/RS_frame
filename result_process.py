import pickle
import torch
import os
import random
import openpyxl
import pandas as pd
from datetime import datetime
from openpyxl.styles import Alignment


def result_process(args, train_epoch_index, loss, other_information, test_epoch_index, all_recall, all_ndcg
                   ,current_best_result):

    #训练情况
    train_result = pd.DataFrame({
        "epoch_index": train_epoch_index,
        "train_loss": loss,
    })
    other_information = torch.stack(other_information, dim=0)
    for i in range(other_information.shape[1]):
        train_result[f"other_information{i + 1}"] = other_information.numpy()[:, i].tolist()

    #测试情况
    test_result = pd.DataFrame({
        "epoch_index": test_epoch_index
    })
    for key, value in all_recall.items():
        test_result[f"{key}"] = value
    for key, value in all_ndcg.items():
        test_result[f"{key}"] = value

    # 获取当前指令
    configs = [i for i in vars(args).items()]
    current_instrument = "python main.py"
    for value in configs:
        current_instrument += f"--{value[0]} {value[1]} "

    #包装结果
    all_result = {
        "current_instrument":current_instrument,
        "train_result":train_result,
        "test_result":test_result,
    }
    for key,value in current_best_result.items():
        all_result[key] = value

    #获取当前时间
    current_time = datetime.now()
    # 格式化输出
    formatted_time = current_time.strftime("%Y-%m-%d_%H;%M;%S")
    #存储路径
    output_file_path = f'./result/{args.model_name}/all_results'
    if not os.path.exists(output_file_path):
        os.makedirs(output_file_path)
    path = f"{output_file_path}/{args.model_name}_{args.dataset_name}_{formatted_time}_{random.randint(1, 10000)}.pkl"
    #存结果
    with open(path, "wb") as file:
        pickle.dump(all_result, file)


def get_best_result(args,path):
        # 获取所有 .pkl 文件的路径
        pkl_files = load_pkl_files(path)
        # 从这些 .pkl 文件中加载数据
        loaded_data = load_data_from_pkl_files(pkl_files)

        for topk in args.topks:
            best_result = loaded_data[0]
            for ele in loaded_data:
                if ele[f"current_best_ndcg{topk}"]>= best_result[f"current_best_ndcg{topk}"] and ele[f"current_best_recall{topk}"]>= best_result[f"current_best_recall{topk}"]:
                    best_result = ele

            #存储路径
            output_file_path = f'./result/{args.model_name}'
            if not os.path.exists(output_file_path):
                os.makedirs(output_file_path)

            path = f"{output_file_path}/{args.model_name}_{args.dataset_name}_topk{topk}_best_result.xlsx"
            with pd.ExcelWriter(path,
                                engine='openpyxl') as writer:
                best_result["train_result"].to_excel(writer, sheet_name='Sheet1', index=False)
                best_result["test_result"].to_excel(writer, sheet_name='Sheet2', index=False)
                # 获取工作簿对象
                workbook = writer.book
                for name in workbook.sheetnames: #指定表格数
                    worksheet = workbook[name]
                    # 设置列宽
                    for idx in range(max(best_result["train_result"].shape[1],best_result["test_result"].shape[1])):
                        column_letter = chr(65 + idx)  # 将列索引转换为列字母
                        worksheet.column_dimensions[column_letter].width = 30
                    # 设置单元格居中对齐
                    for row in worksheet.iter_rows():
                        for cell in row:
                            cell.alignment = Alignment(horizontal='center', vertical='center')

            # 加载现有的 Excel 文件
            wb = openpyxl.load_workbook(path)
            # 添加一个新的工作表
            ws = wb.create_sheet(title="Sheet0", index=0)  # index=0 表示将新工作表放在最前面
            # 将字符串写入新工作表的第一行的前四个单元格
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(best_result["current_instrument"])//2)
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(best_result["current_instrument"]) // 2)
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(best_result["current_instrument"]) // 2)
            ws.cell(row=1, column=1).value = best_result["current_instrument"]
            ndcg = round(best_result[f"current_best_ndcg{topk}"],4)
            recall = round(best_result[f"current_best_recall{topk}"],4)
            ws.cell(row=2,column=1).value = f"best_recall{topk}:{recall}"
            ws.cell(row=3,column=1).value = f"best_ndcg{topk}:{ndcg}"
            wb.save(path)# 保存修改后的 Excel 文件

            #存字典信息
            save_weight_path = f'./saved_model/{args.model_name}'
            if not os.path.exists(save_weight_path):
                os.makedirs(save_weight_path)
            torch.save(best_result[f"current_best_{topk}_state_dict"], save_weight_path+f'/saved_best_state_dict_{args.model_name}_{args.dataset_name}_topk{topk}.pt')


def load_pkl_files(directory):
    pkl_files = []
    # 遍历指定目录下的所有文件
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith(".pkl"):
                pkl_files.append(os.path.join(root, file))

    return pkl_files


def load_data_from_pkl_files(pkl_files):
    data_list = []
    for file_path in pkl_files:
        with open(file_path, "rb") as file:
            data = pickle.load(file)
            data_list.append(data)
    return data_list
